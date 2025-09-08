from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from groups.models import Group
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .utils import get_buyer_list_data
import urllib.parse
from django.utils import timezone


def export_orders_to_excel(request, group_id):
    group = get_object_or_404(Group, id=group_id, owner=request.user)

    if group.status != "reached":
        messages.error(request, "只有已成團團購可以匯出訂單")
        return redirect("orders:buyer_list", group_id)

    group, orders, buyers = get_buyer_list_data(request.user, group_id, for_export=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "跟團者資訊"

    border_color = "e1e1e1"
    border_style = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    headers = [
        "訂單號碼",
        "收件者",
        "電話",
        "郵遞區號",
        "地址",
        "品項",
        "數量",
        "單價",
        "小計",
        "合計",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(vertical="center")
        cell.border = border_style

    current_row = 2

    for order_idx, order in enumerate(orders):
        fill_color = "F2F2F2" if order_idx % 2 == 0 else None

        start_row = current_row
        jgps = order.joined_group.joined_group_products.all()

        for jgp in jgps:
            contents = [
                order.order_number,
                order.ship_recipient_name,
                order.ship_phone,
                order.ship_postal_code,
                f"{order.ship_county}{order.ship_district}{order.ship_road}{order.ship_detail}",
                jgp.product.name,
                jgp.quantity,
                jgp.product.price,
                jgp.subtotal,
                order.amount,
            ]
            for col, content in enumerate(contents, 1):
                cell = ws.cell(row=current_row, column=col, value=content)
                cell.alignment = Alignment(vertical="center")

                if col in [8, 9, 10]:
                    cell.number_format = '"$"#,##0'

                if fill_color:
                    cell.fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
                cell.border = border_style

            current_row += 1

        end_row = current_row - 1
        if len(jgps) > 1:
            merge_cols = [1, 2, 3, 4, 5, 10]
            for col in merge_cols:
                col_letter = get_column_letter(col)
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(10, min(max_length + 2, 50))
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.column_dimensions['E'].width = 35  # 地址欄位加寬
    ws.column_dimensions['F'].width = 30  # 品項欄位加寬

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = urllib.parse.quote(
        f"{group.name}－跟團者資訊_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"

    wb.save(response)
    return response
